import streamlit as st
import openpyxl
import xml.etree.ElementTree as ET
import csv
import re
import tempfile
import os
import uuid
import zipfile
from io import StringIO

# --- UI Setup ---
st.set_page_config(page_title="Missing Strings Report", page_icon="🔍", layout="wide")

st.markdown(
    """
    <div style='text-align: right; color: gray; font-size: 14px; margin-bottom: -20px;'>
        <i>developed by Ahmet Ozerdem</i>
    </div>
    """, 
    unsafe_allow_html=True
)

st.title("🔍 Missing Strings Report Generator")
st.markdown("Identifies completely skipped cells and improperly segmented (split) strings. Supports single file pairs or bulk ZIP uploads.")

# --- CORE LQA LOGIC FUNCTION ---
def process_file_pair(xlsx_path, mxliff_path, file_label):
    """Processes a single XLSX and MXLIFF pair and returns a list of missing/split items."""
    tree = ET.parse(mxliff_path)
    root = tree.getroot()
    
    para_blocks = {}
    
    # 1. Group all MXLIFF segments by their native Memsource Paragraph ID
    for tu in root.iter('{urn:oasis:names:tc:xliff:document:1.2}trans-unit'):
        para_id = None
        for key, val in tu.attrib.items():
            if key.endswith('para-id'):
                para_id = val
                break
                
        if not para_id:
            para_id = str(uuid.uuid4())
            
        source_node = tu.find('{urn:oasis:names:tc:xliff:document:1.2}source')
        if source_node is not None:
            text = "".join(source_node.itertext())
            if para_id not in para_blocks:
                para_blocks[para_id] = []
            para_blocks[para_id].append(text)

    # 2. Create two buckets: Normal Strings and Split Strings
    normal_mxliff_pool = []
    split_mxliff_pool = []
    
    for pid, texts in para_blocks.items():
        combined_text = "".join(texts)
        cleaned = re.sub(r'[\s\W_]+', '', combined_text)
        if cleaned:
            if len(texts) == 1:
                normal_mxliff_pool.append(cleaned)
            else:
                split_mxliff_pool.append(cleaned)

    # Parse the source XLSX file
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active 
    
    file_items = []
    target_columns = ['F', 'G', 'H', 'I', 'J', 'K']
    
    # 3. Iterate through Excel and categorize
    for row in range(2, sheet.max_row + 1):
        for col_index, col_letter in enumerate(target_columns, start=6):
            cell = sheet.cell(row=row, column=col_index)
            
            if cell.value is not None:
                cell_text = str(cell.value).strip()
                
                if cell_text: 
                    cleaned_cell = re.sub(r'[\s\W_]+', '', cell_text)
                    
                    if not cleaned_cell:
                        continue 
                        
                    # Check 1: Normal 1:1 match
                    if cleaned_cell in normal_mxliff_pool:
                        normal_mxliff_pool.remove(cleaned_cell)
                        
                    # Check 2: ATMS split string
                    elif cleaned_cell in split_mxliff_pool:
                        split_mxliff_pool.remove(cleaned_cell)
                        file_items.append({
                            'File': file_label,
                            'Cell': f"{col_letter}{row}",
                            'Status': 'Split String', # Emojis removed for clean CSV
                            'Source Text': cell_text
                        })
                        
                    # Check 3: Genuinely missing
                    else:
                        file_items.append({
                            'File': file_label,
                            'Cell': f"{col_letter}{row}",
                            'Status': 'Missing Cell', # Emojis removed for clean CSV
                            'Source Text': cell_text
                        })
                        
    return file_items

# --- UI TABS ---
tab_single, tab_batch = st.tabs(["📄 Single File Upload", "🗂️ Batch Upload (ZIP)"])

report_items = None
dynamic_filename = "LQA_Report.csv"

# ==========================================
# MODE 1: SINGLE FILE UPLOAD
# ==========================================
with tab_single:
    col1, col2 = st.columns(2)
    with col1:
        xlsx_file = st.file_uploader("Upload Source XLSX File", type=['xlsx'], key="single_x")
    with col2:
        mxliff_file = st.file_uploader("Upload Processed MXLIFF File", type=['mxliff'], key="single_m")

    if xlsx_file and mxliff_file:
        if st.button("Generate Report", use_container_width=True, key="btn_single"):
            with st.spinner("Processing files..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
                    tmp_xlsx.write(xlsx_file.getvalue())
                    xlsx_path = tmp_xlsx.name
                    
                with tempfile.NamedTemporaryFile(delete=False, suffix=".mxliff") as tmp_mxliff:
                    tmp_mxliff.write(mxliff_file.getvalue())
                    mxliff_path = tmp_mxliff.name

                try:
                    base_name = xlsx_file.name.replace('.xlsx', '')
                    dynamic_filename = f"LQA_Report_{base_name[:15]}.csv"
                    report_items = process_file_pair(xlsx_path, mxliff_path, base_name)
                except Exception as e:
                    st.error(f"An error occurred during processing:\n{e}")
                finally:
                    if os.path.exists(xlsx_path): os.remove(xlsx_path)
                    if os.path.exists(mxliff_path): os.remove(mxliff_path)

# ==========================================
# MODE 2: BATCH ZIP UPLOAD
# ==========================================
with tab_batch:
    st.info("Upload a ZIP file containing your XLSX and MXLIFF files. The tool will automatically match them based on their filenames.")
    zip_file = st.file_uploader("Upload ZIP Archive", type=['zip'], key="batch_zip")
    
    if zip_file:
        if st.button("Generate Consolidated Report", use_container_width=True, key="btn_batch"):
            with st.spinner("Extracting and mapping files..."):
                report_items = []
                dynamic_filename = "LQA_Consolidated_Batch_Report.csv"
                
                with tempfile.TemporaryDirectory() as temp_dir:
                    # 1. Extract ZIP
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                        
                    # 2. Find and categorize files
                    xlsx_files = []
                    mxliff_files = []
                    
                    for root_dir, dirs, files in os.walk(temp_dir):
                        for file in files:
                            if "__MACOSX" in root_dir or file.startswith("._"):
                                continue # Ignore Mac hidden files
                                
                            full_path = os.path.join(root_dir, file)
                            if file.endswith('.xlsx'):
                                xlsx_files.append((file, full_path))
                            elif file.endswith('.mxliff'):
                                mxliff_files.append((file, full_path))
                                
                    # 3. Map pairs and process
                    unmatched_xlsx = []
                    processed_count = 0
                    
                    for x_file, x_path in xlsx_files:
                        base_name = x_file.replace('.xlsx', '')
                        matched_m_path = None
                        
                        # Find matching MXLIFF by checking if it starts with the XLSX base name
                        for m_file, m_path in mxliff_files:
                            if m_file.startswith(base_name):
                                matched_m_path = m_path
                                break
                                
                        if matched_m_path:
                            try:
                                items = process_file_pair(x_path, matched_m_path, base_name)
                                report_items.extend(items)
                                processed_count += 1
                            except Exception as e:
                                st.error(f"Error processing {x_file}: {e}")
                        else:
                            unmatched_xlsx.append(x_file)
                            
                    st.success(f"Successfully processed {processed_count} file pairs!")
                    if unmatched_xlsx:
                        st.warning(f"Could not find matching MXLIFF files for: {', '.join(unmatched_xlsx)}")

# ==========================================
# SHARED OUTPUT & EXPORT LOGIC
# ==========================================
if report_items is not None: # Triggers if either button was pressed and finished
    missing_only = [item for item in report_items if item['Status'] == 'Missing Cell']
    split_only = [item for item in report_items if item['Status'] == 'Split String']
    
    if not report_items:
        st.success("✅ No hidden strings or segmentation issues found in the processed files! Everything looks perfect.")
    else:
        st.divider()
        st.markdown("### 📊 Scan Results")
        metric_col1, metric_col2 = st.columns(2)
        metric_col1.metric("Completely Missing Cells", len(missing_only))
        metric_col2.metric("Improperly Split Strings", len(split_only))
        
        st.markdown("### Consolidated Findings Preview")
        st.dataframe(report_items, use_container_width=True)
        
        csv_buffer = StringIO()
        writer = csv.DictWriter(csv_buffer, fieldnames=['File', 'Cell', 'Status', 'Source Text'])
        writer.writeheader()
        writer.writerows(report_items)
        
        st.download_button(
            label="⬇️ Download Full CSV Report",
            data=csv_buffer.getvalue(),
            file_name=dynamic_filename,
            mime="text/csv",
            type="primary"
        )
